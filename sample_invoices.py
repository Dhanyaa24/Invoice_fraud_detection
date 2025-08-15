import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta
import random

def create_sample_excel_invoices():
    """Create sample Excel invoice files for testing fraud detection"""
    
    # Sample 1: Low Risk Invoice (Legitimate)
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Invoice"
    
    # Header
    ws1['A1'] = "INVOICE"
    ws1['A1'].font = Font(size=16, bold=True)
    ws1.merge_cells('A1:D1')
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    # Invoice details
    ws1['A3'] = "Invoice Number:"
    ws1['B3'] = "INV-2024-001"
    ws1['A4'] = "Date:"
    ws1['B4'] = "2024-08-10"
    ws1['A5'] = "Due Date:"
    ws1['B5'] = "2024-09-10"
    
    ws1['C3'] = "Vendor:"
    ws1['D3'] = "Beta Traders"
    ws1['C4'] = "Contact:"
    ws1['D4'] = "john@betatraders.com"
    ws1['C5'] = "Phone:"
    ws1['D5'] = "+1-555-0123"
    
    # Items
    ws1['A7'] = "Item"
    ws1['B7'] = "Description"
    ws1['C7'] = "Quantity"
    ws1['D7'] = "Unit Price"
    ws1['E7'] = "Total"
    
    # Style headers
    for cell in ['A7', 'B7', 'C7', 'D7', 'E7']:
        ws1[cell].font = Font(bold=True)
        ws1[cell].alignment = Alignment(horizontal='center')
    
    # Sample items
    items = [
        ["Office Supplies", "Premium pens and notebooks", 50, 2.45, 122.50],
        ["Paper", "A4 printer paper, 500 sheets", 10, 8.99, 89.90],
        ["Desk Organizer", "Multi-compartment organizer", 5, 15.75, 78.75]
    ]
    
    for i, item in enumerate(items, 8):
        ws1[f'A{i}'] = item[0]
        ws1[f'B{i}'] = item[1]
        ws1[f'C{i}'] = item[2]
        ws1[f'D{i}'] = f"${item[3]:.2f}"
        ws1[f'E{i}'] = f"${item[4]:.2f}"
    
    # Total
    total = sum(item[4] for item in items)
    ws1[f'A{8+len(items)}'] = "TOTAL:"
    ws1[f'E{8+len(items)}'] = f"${total:.2f}"
    ws1[f'A{8+len(items)}'].font = Font(bold=True)
    ws1[f'E{8+len(items)}'].font = Font(bold=True)
    
    wb1.save("sample_invoice_low_risk.xlsx")
    print(f"Created: sample_invoice_low_risk.xlsx (Amount: ${total:.2f})")
    
    # Sample 2: Medium Risk Invoice (Suspicious)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Invoice"
    
    ws2['A1'] = "INVOICE"
    ws2['A1'].font = Font(size=16, bold=True)
    ws2.merge_cells('A1:D1')
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    ws2['A3'] = "Invoice Number:"
    ws2['B3'] = "INV-2024-002"
    ws2['A4'] = "Date:"
    ws2['B4'] = "2024-08-10"
    ws2['A5'] = "Due Date:"
    ws2['B5'] = "2024-09-10"
    
    ws2['C3'] = "Vendor:"
    ws2['D3'] = "Gamma Co"
    ws2['C4'] = "Contact:"
    ws2['D4'] = "sales@gammaco.com"
    ws2['C5'] = "Phone:"
    ws2['D5'] = "+1-555-0456"
    
    ws2['A7'] = "Item"
    ws2['B7'] = "Description"
    ws2['C7'] = "Quantity"
    ws2['D7'] = "Unit Price"
    ws2['E7'] = "Total"
    
    for cell in ['A7', 'B7', 'C7', 'D7', 'E7']:
        ws2[cell].font = Font(bold=True)
        ws2[cell].alignment = Alignment(horizontal='center')
    
    items2 = [
        ["Consulting Services", "Business process optimization", 40, 125.00, 5000.00]
    ]
    
    for i, item in enumerate(items2, 8):
        ws2[f'A{i}'] = item[0]
        ws2[f'B{i}'] = item[1]
        ws2[f'C{i}'] = item[2]
        ws2[f'D{i}'] = f"${item[3]:.2f}"
        ws2[f'E{i}'] = f"${item[4]:.2f}"
    
    total2 = sum(item[4] for item in items2)
    ws2[f'A{8+len(items2)}'] = "TOTAL:"
    ws2[f'E{8+len(items2)}'] = f"${total2:.2f}"
    ws2[f'A{8+len(items2)}'].font = Font(bold=True)
    ws2[f'E{8+len(items2)}'].font = Font(bold=True)
    
    wb2.save("sample_invoice_medium_risk.xlsx")
    print(f"Created: sample_invoice_medium_risk.xlsx (Amount: ${total2:.2f})")
    
    # Sample 3: High Risk Invoice (Very Suspicious)
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Invoice"
    
    ws3['A1'] = "INVOICE"
    ws3['A1'].font = Font(size=16, bold=True)
    ws3.merge_cells('A1:D1')
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    ws3['A3'] = "Invoice Number:"
    ws3['B3'] = "INV-2024-003"
    ws3['A4'] = "Date:"
    ws3['B4'] = "2024-08-10"
    ws3['A5'] = "Due Date:"
    ws3['B5'] = "2024-09-10"
    
    ws3['C3'] = "Vendor:"
    ws3['D3'] = "Delta Services"
    ws3['C4'] = "Contact:"
    ws3['D4'] = "billing@deltaservices.com"
    ws3['C5'] = "Phone:"
    ws3['D5'] = "+1-555-0789"
    
    ws3['A7'] = "Item"
    ws3['B7'] = "Description"
    ws3['C7'] = "Quantity"
    ws3['D7'] = "Unit Price"
    ws3['E7'] = "Total"
    
    for cell in ['A7', 'B7', 'C7', 'D7', 'E7']:
        ws3[cell].font = Font(bold=True)
        ws3[cell].alignment = Alignment(horizontal='center')
    
    items3 = [
        ["Premium Services", "Executive consulting package", 1, 10000.00, 10000.00]
    ]
    
    for i, item in enumerate(items3, 8):
        ws3[f'A{i}'] = item[0]
        ws3[f'B{i}'] = item[1]
        ws3[f'C{i}'] = item[2]
        ws3[f'D{i}'] = f"${item[3]:.2f}"
        ws3[f'E{i}'] = f"${item[4]:.2f}"
    
    total3 = sum(item[4] for item in items3)
    ws3[f'A{8+len(items2)}'] = "TOTAL:"
    ws3[f'E{8+len(items2)}'] = f"${total3:.2f}"
    ws3[f'A{8+len(items2)}'].font = Font(bold=True)
    ws3[f'E{8+len(items2)}'].font = Font(bold=True)
    
    wb3.save("sample_invoice_high_risk.xlsx")
    print(f"Created: sample_invoice_high_risk.xlsx (Amount: ${total3:.2f})")

if __name__ == "__main__":
    create_sample_excel_invoices()
    print("\nSample Excel invoices created successfully!")
    print("Upload these files to test different fraud detection scenarios:")
    print("- sample_invoice_low_risk.xlsx: Small amount, detailed items")
    print("- sample_invoice_medium_risk.xlsx: Round amount, consulting services")
    print("- sample_invoice_high_risk.xlsx: Very large round amount") 